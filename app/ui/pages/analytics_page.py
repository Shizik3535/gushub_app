from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QLabel, QPushButton, QFrame,
                            QHBoxLayout, QLineEdit, QTableWidget, QTableWidgetItem,
                            QStackedWidget, QHeaderView, QGridLayout, QFileDialog,
                            QMessageBox)
from PyQt6.QtCore import Qt, pyqtSignal
from app.api.gushub_api import GushubAPI
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import openpyxl.cell.cell

class StudentsListWidget(QWidget):
    """Виджет для отображения списка студентов"""
    student_selected = pyqtSignal(int)  # Сигнал с ID выбранного студента
    back_clicked = pyqtSignal()  # Сигнал для возврата к выбору раздела
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.gushub_api = GushubAPI()
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Добавляем поиск
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск студента...")
        self.search_input.textChanged.connect(self.filter_students)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)
        
        # Добавляем таблицу студентов
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Имя пользователя", "Полное имя"])
        self.table.verticalHeader().setVisible(False)  # Скрываем боковую нумерацию
        self.table.horizontalHeader().setStretchLastSection(True)  # Растягиваем последнюю колонку
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Растягиваем первую колонку
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Растягиваем вторую колонку
        self.table.cellClicked.connect(self.on_student_selected)
        layout.addWidget(self.table)
        
        # Добавляем кнопку возврата
        back_button = QPushButton("← Вернуться к выбору раздела")
        back_button.clicked.connect(self.back_clicked.emit)
        layout.addWidget(back_button)
    
    def load_data(self):
        """Загрузка данных о студентах"""
        try:
            users = self.gushub_api.get_users()
            self.table.setRowCount(len(users))
            for i, user in enumerate(users):
                self.table.setItem(i, 0, QTableWidgetItem(user.username))
                self.table.setItem(i, 1, QTableWidgetItem(f"{user.firstName} {user.lastName}".strip() or user.username))
        except Exception as e:
            pass
    
    def filter_students(self, text: str):
        """Фильтрация таблицы студентов"""
        for row in range(self.table.rowCount()):
            show_row = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text.lower() in item.text().lower():
                    show_row = True
                    break
            self.table.setRowHidden(row, not show_row)
    
    def on_student_selected(self, row: int, column: int):
        """Обработка выбора студента"""
        try:
            users = self.gushub_api.get_users()
            student_id = users[row].id
            self.student_selected.emit(student_id)
        except Exception as e:
            pass

class GroupsListWidget(QWidget):
    """Виджет для отображения списка групп"""
    group_selected = pyqtSignal(int)  # Сигнал с ID выбранной группы
    back_clicked = pyqtSignal()  # Сигнал для возврата к выбору раздела
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.gushub_api = GushubAPI()
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Добавляем поиск
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск группы...")
        self.search_input.textChanged.connect(self.filter_groups)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)
        
        # Добавляем таблицу групп
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Название", "Описание"])
        self.table.verticalHeader().setVisible(False)  # Скрываем боковую нумерацию
        self.table.horizontalHeader().setStretchLastSection(True)  # Растягиваем последнюю колонку
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Растягиваем первую колонку
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Растягиваем вторую колонку
        self.table.cellClicked.connect(self.on_group_selected)
        layout.addWidget(self.table)
        
        # Добавляем кнопку возврата
        back_button = QPushButton("← Вернуться к выбору раздела")
        back_button.clicked.connect(self.back_clicked.emit)
        layout.addWidget(back_button)
    
    def load_data(self):
        """Загрузка данных о группах"""
        try:
            groups = self.gushub_api.get_groups()
            self.table.setRowCount(len(groups))
            for i, group in enumerate(groups):
                self.table.setItem(i, 0, QTableWidgetItem(group.name))
                self.table.setItem(i, 1, QTableWidgetItem(group.description))
        except Exception as e:
            pass
    
    def filter_groups(self, text: str):
        """Фильтрация таблицы групп"""
        for row in range(self.table.rowCount()):
            show_row = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text.lower() in item.text().lower():
                    show_row = True
                    break
            self.table.setRowHidden(row, not show_row)
    
    def on_group_selected(self, row: int, column: int):
        """Обработка выбора группы"""
        try:
            groups = self.gushub_api.get_groups()
            group_id = groups[row].id
            self.group_selected.emit(group_id)
        except Exception as e:
            pass

class StudentStatsWidget(QWidget):
    """Виджет для отображения статистики студента"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.gushub_api = GushubAPI()
        self.setup_ui()

    def localize(self, value):
        return value if value not in (None, '', 'None') else 'Не задано'
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Добавляем заголовок и кнопку возврата
        header_layout = QHBoxLayout()
        self.title = QLabel("<h2>Статистика студента</h2>")
        self.back_button = QPushButton("← Назад к списку")
        header_layout.addWidget(self.title)
        header_layout.addWidget(self.back_button)
        layout.addLayout(header_layout)
        
        # Создаем сетку для карточек
        grid_layout = QGridLayout()
        grid_layout.setSpacing(5)
        
        # Карточка основной информации
        info_frame = QFrame()
        info_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        info_frame.setMinimumHeight(200)
        info_layout = QVBoxLayout(info_frame)
        info_title = QLabel("<h3>Основная информация</h3>")
        info_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        info_layout.addWidget(info_title)
        self.info_label = QLabel()
        self.info_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        info_layout.addWidget(self.info_label)
        info_layout.addStretch(0)
        grid_layout.addWidget(info_frame, 0, 0)
        
        # Карточка курсов
        courses_frame = QFrame()
        courses_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        courses_frame.setMinimumHeight(200)
        courses_layout = QVBoxLayout(courses_frame)
        courses_title = QLabel("<h3>Статистика по курсам</h3>")
        courses_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        courses_layout.addWidget(courses_title)
        self.courses_label = QLabel()
        self.courses_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        courses_layout.addWidget(self.courses_label)
        courses_layout.addStretch(0)
        grid_layout.addWidget(courses_frame, 0, 1)
        
        # Карточка заданий
        tasks_frame = QFrame()
        tasks_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        tasks_frame.setMinimumHeight(200)
        tasks_layout = QVBoxLayout(tasks_frame)
        tasks_title = QLabel("<h3>Статистика по заданиям</h3>")
        tasks_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        tasks_layout.addWidget(tasks_title)
        self.tasks_label = QLabel()
        self.tasks_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        tasks_layout.addWidget(self.tasks_label)
        tasks_layout.addStretch(0)
        grid_layout.addWidget(tasks_frame, 1, 0)
        
        # Карточка оценок
        grades_frame = QFrame()
        grades_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        grades_frame.setMinimumHeight(200)
        grades_layout = QVBoxLayout(grades_frame)
        grades_title = QLabel("<h3>Статистика по оценкам</h3>")
        grades_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        grades_layout.addWidget(grades_title)
        self.grades_label = QLabel()
        self.grades_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        grades_layout.addWidget(self.grades_label)
        grades_layout.addStretch(0)
        grid_layout.addWidget(grades_frame, 1, 1)
        
        layout.addLayout(grid_layout)

        # Кнопка экспорта
        export_layout = QHBoxLayout()
        self.export_button = QPushButton("Экспортировать в Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        export_layout.addWidget(self.export_button)
        layout.addLayout(export_layout)

    def export_to_excel(self):
        try:
            # Получаем данные студента
            user = self.gushub_api.get_user(self.student_id)
            if not user:
                raise Exception("Не удалось получить данные пользователя")

            stats = self.gushub_api.get_user_statistics(self.student_id)
            if not stats:
                raise Exception("Не удалось получить статистику пользователя")

            grades_stats = self.gushub_api.get_user_grades_statistics(self.student_id)
            if not grades_stats:
                raise Exception("Не удалось получить статистику оценок")

            # Создаем новый Excel-файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Статистика студента"

            # Определяем стиль обводки
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            try:
                # Заголовок
                ws['A1'] = f"Статистика студента: {self.localize(user.username)}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:B1')
                ws['A1'].border = thin_border
                ws['B1'].border = thin_border

                # Основная информация
                ws['A3'] = "Основная информация"
                ws['A3'].font = Font(bold=True)
                ws['A3'].border = thin_border
                ws['B3'].border = thin_border
                ws['A4'] = "Имя:"
                ws['B4'] = self.localize(user.firstName)
                ws['A5'] = "Email:"
                ws['B5'] = self.localize(user.email)
                ws['A6'] = "Роль:"
                ws['B6'] = self.localize(user.role)
                ws['A7'] = "Дата регистрации:"
                ws['B7'] = user.createdAt.strftime('%d.%m.%Y %H:%M') if user.createdAt else 'Не задано'

                # Статистика по курсам
                ws['A9'] = "Статистика по курсам"
                ws['A9'].font = Font(bold=True)
                ws['A9'].border = thin_border
                ws['B9'].border = thin_border
                ws['A10'] = "Всего курсов:"
                ws['B10'] = self.localize(stats.totalCourses)
                ws['A11'] = "Завершено курсов:"
                ws['B11'] = self.localize(stats.completedCourses)
                ws['A12'] = "В процессе:"
                ws['B12'] = self.localize(stats.totalCourses - stats.completedCourses if stats.totalCourses is not None and stats.completedCourses is not None else None)
                ws['A13'] = "Общее время обучения:"
                ws['B13'] = f"{self.localize(stats.totalTimeSpent)} минут"

                # Статистика по заданиям
                ws['A15'] = "Статистика по заданиям"
                ws['A15'].font = Font(bold=True)
                ws['A15'].border = thin_border
                ws['B15'].border = thin_border
                ws['A16'] = "Всего заданий:"
                ws['B16'] = self.localize(stats.totalTasks)
                ws['A17'] = "Выполнено заданий:"
                ws['B17'] = self.localize(stats.completedTasks)
                ws['A18'] = "В процессе:"
                ws['B18'] = self.localize(stats.totalTasks - stats.completedTasks if stats.totalTasks is not None and stats.completedTasks is not None else None)
                ws['A19'] = "Процент выполнения:"
                ws['B19'] = f"{(stats.completedTasks / stats.totalTasks * 100) if stats.totalTasks else 0:.1f}%"

                # Статистика по оценкам
                ws['A21'] = "Статистика по оценкам"
                ws['A21'].font = Font(bold=True)
                ws['A21'].border = thin_border
                ws['B21'].border = thin_border
                ws['A22'] = "Средний балл:"
                ws['B22'] = self.localize(f'{grades_stats.averageGrade:.1f}' if grades_stats.averageGrade is not None else None)
                ws['A23'] = "Всего оценок:"
                ws['B23'] = self.localize(grades_stats.totalGrades)
                ws['A24'] = "Распределение оценок:"
                ws['A24'].font = Font(bold=True)
                ws['A24'].border = thin_border
                ws['B24'].border = thin_border
                ws['A25'] = "2:"
                ws['B25'] = self.localize(grades_stats.gradesByValue.get('2'))
                ws['A26'] = "3:"
                ws['B26'] = self.localize(grades_stats.gradesByValue.get('3'))
                ws['A27'] = "4:"
                ws['B27'] = self.localize(grades_stats.gradesByValue.get('4'))
                ws['A28'] = "5:"
                ws['B28'] = self.localize(grades_stats.gradesByValue.get('5'))

                # Добавляем обводку для всех ячеек с данными
                for row in range(4, 29):
                    if row not in [8, 14, 20]:
                        for col in ['A', 'B']:
                            cell = ws[f'{col}{row}']
                            cell.border = thin_border

                # Выравнивание
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                # Автоматическая настройка ширины столбцов
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        try:
                            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                if column_letter is None:
                                    column_letter = cell.column_letter
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                        except:
                            pass
                    if column_letter:
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column_letter].width = adjusted_width

                # Объединяем ячейки для заголовков разделов
                ws.merge_cells('A1:B1')  # Заголовок
                ws.merge_cells('A3:B3')  # Основная информация
                ws.merge_cells('A9:B9')  # Статистика по курсам
                ws.merge_cells('A15:B15')  # Статистика по заданиям
                ws.merge_cells('A21:B21')  # Статистика по оценкам
                ws.merge_cells('A24:B24')  # Распределение оценок

                # Сохраняем файл
                file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Excel Files (*.xlsx)")
                if file_path:
                    wb.save(file_path)
                    QMessageBox.information(self, "Успех", "Статистика успешно экспортирована в Excel!")

            except Exception as e:
                raise

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при экспорте в Excel:\n{str(e)}")

    def load_data(self, student_id: int):
        self.student_id = student_id  # Сохраняем ID студента для экспорта
        try:
            user = self.gushub_api.get_user(student_id)
            stats = self.gushub_api.get_user_statistics(student_id)
            grades_stats = self.gushub_api.get_user_grades_statistics(student_id)
            self.title.setText(f"<h2>Статистика студента: {self.localize(user.username)}</h2>")

            # Основная информация
            info_text = f"""
            <table style='margin-top:0;'>
                <tr><td><b>Имя:</b></td><td>{self.localize(user.firstName)}</td></tr>
                <tr><td><b>Email:</b></td><td>{self.localize(user.email)}</td></tr>
                <tr><td><b>Роль:</b></td><td>{self.localize(user.role)}</td></tr>
                <tr><td><b>Дата регистрации:</b></td><td>{user.createdAt.strftime('%d.%m.%Y %H:%M') if user.createdAt else 'Не задано'}</td></tr>
            </table>
            """
            self.info_label.setText(info_text)

            # Курсы
            courses_text = f"""
            <table style='margin-top:0;'>
                <tr><td><b>Всего курсов:</b></td><td>{self.localize(stats.totalCourses)}</td></tr>
                <tr><td><b>Завершено курсов:</b></td><td>{self.localize(stats.completedCourses)}</td></tr>
                <tr><td><b>В процессе:</b></td><td>{self.localize(stats.totalCourses - stats.completedCourses if stats.totalCourses is not None and stats.completedCourses is not None else None)}</td></tr>
                <tr><td><b>Общее время обучения:</b></td><td>{self.localize(stats.totalTimeSpent)} минут</td></tr>
            </table>
            """
            self.courses_label.setText(courses_text)

            # Задания
            percent = (stats.completedTasks / stats.totalTasks * 100) if stats.totalTasks else 0
            tasks_text = f"""
            <table style='margin-top:0;'>
                <tr><td><b>Всего заданий:</b></td><td>{self.localize(stats.totalTasks)}</td></tr>
                <tr><td><b>Выполнено заданий:</b></td><td>{self.localize(stats.completedTasks)}</td></tr>
                <tr><td><b>В процессе:</b></td><td>{self.localize(stats.totalTasks - stats.completedTasks if stats.totalTasks is not None and stats.completedTasks is not None else None)}</td></tr>
                <tr><td><b>Процент выполнения:</b></td><td>{percent:.1f}%</td></tr>
            </table>
            """
            self.tasks_label.setText(tasks_text)

            # Оценки
            grades_text = f"""
            <table style='margin-top:0;'>
                <tr><td><b>Средний балл:</b></td><td>{self.localize(f'{grades_stats.averageGrade:.1f}' if grades_stats.averageGrade is not None else None)}</td></tr>
                <tr><td><b>Всего оценок:</b></td><td>{self.localize(grades_stats.totalGrades)}</td></tr>
                <tr><td colspan='2'><b>Распределение оценок:</b></td></tr>
                <tr><td>2:</td><td>{self.localize(grades_stats.gradesByValue.get('2'))}</td></tr>
                <tr><td>3:</td><td>{self.localize(grades_stats.gradesByValue.get('3'))}</td></tr>
                <tr><td>4:</td><td>{self.localize(grades_stats.gradesByValue.get('4'))}</td></tr>
                <tr><td>5:</td><td>{self.localize(grades_stats.gradesByValue.get('5'))}</td></tr>
            </table>
            """
            self.grades_label.setText(grades_text)

        except Exception as e:
            self.info_label.setText("Ошибка при загрузке данных")
            self.courses_label.setText("")
            self.tasks_label.setText("")
            self.grades_label.setText("")

class GroupStatsWidget(QWidget):
    """Виджет для отображения статистики группы"""
    student_selected = pyqtSignal(int)  # Сигнал с ID выбранного студента

    def __init__(self, parent=None):
        super().__init__(parent)
        self.gushub_api = GushubAPI()
        self.setup_ui()

    def localize(self, value):
        return value if value not in (None, '', 'None') else 'Не задано'
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Заголовок и кнопка возврата
        header_layout = QHBoxLayout()
        self.title = QLabel("<h2>Статистика группы</h2>")
        self.back_button = QPushButton("← Назад к списку")
        header_layout.addWidget(self.title)
        header_layout.addWidget(self.back_button)
        layout.addLayout(header_layout)

        # Сетка для карточек
        grid_layout = QGridLayout()
        grid_layout.setSpacing(5)

        # Карточка основной информации
        info_frame = QFrame()
        info_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        info_frame.setMinimumHeight(200)
        info_layout = QVBoxLayout(info_frame)
        info_title = QLabel("<h3>Основная информация</h3>")
        info_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        info_layout.addWidget(info_title)
        self.info_label = QLabel()
        self.info_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        info_layout.addWidget(self.info_label)
        info_layout.addStretch(0)
        grid_layout.addWidget(info_frame, 0, 0)

        # Карточка участников (таблица)
        members_frame = QFrame()
        members_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        members_frame.setMinimumHeight(200)
        members_layout = QVBoxLayout(members_frame)
        members_title = QLabel("<h3>Участники</h3>")
        members_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        members_layout.addWidget(members_title)
        self.members_table = QTableWidget()
        self.members_table.setColumnCount(2)
        self.members_table.setHorizontalHeaderLabels(["Имя", "Логин"])
        self.members_table.verticalHeader().setVisible(False)
        self.members_table.horizontalHeader().setStretchLastSection(True)
        self.members_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.members_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.members_table.cellClicked.connect(self.on_member_selected)
        members_layout.addWidget(self.members_table)
        members_layout.addStretch(0)
        grid_layout.addWidget(members_frame, 0, 1)

        # Карточка курсов
        courses_frame = QFrame()
        courses_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        courses_frame.setMinimumHeight(200)
        courses_layout = QVBoxLayout(courses_frame)
        courses_title = QLabel("<h3>Курсы группы</h3>")
        courses_title.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        courses_layout.addWidget(courses_title)
        self.courses_label = QLabel()
        self.courses_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        courses_layout.addWidget(self.courses_label)
        courses_layout.addStretch(0)
        grid_layout.addWidget(courses_frame, 1, 0, 1, 2)

        layout.addLayout(grid_layout)

        # Кнопка экспорта
        export_layout = QHBoxLayout()
        self.export_button = QPushButton("Экспортировать в Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        export_layout.addWidget(self.export_button)
        layout.addLayout(export_layout)

    def export_to_excel(self):
        try:
            # Получаем данные группы
            group = self.gushub_api.get_group(self.group_id)
            if not group:
                raise Exception("Не удалось получить данные группы")

            # Создаем новый Excel-файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Статистика группы"

            # Определяем стиль обводки
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            try:
                # Заголовок
                ws['A1'] = f"Статистика группы: {self.localize(group.name)}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:B1')
                ws['A1'].border = thin_border
                ws['B1'].border = thin_border

                # Основная информация
                ws['A3'] = "Основная информация"
                ws['A3'].font = Font(bold=True)
                ws['A3'].border = thin_border
                ws['B3'].border = thin_border
                ws['A4'] = "Название:"
                ws['B4'] = self.localize(group.name)
                ws['A4'].border = thin_border
                ws['B4'].border = thin_border
                ws['A5'] = "Описание:"
                ws['B5'] = self.localize(group.description)
                ws['A5'].border = thin_border
                ws['B5'].border = thin_border
                ws['A6'] = "Код приглашения:"
                ws['B6'] = self.localize(group.inviteCode)
                ws['A6'].border = thin_border
                ws['B6'].border = thin_border
                ws['A7'] = "Дата создания:"
                ws['B7'] = group.createdAt.strftime('%d.%m.%Y %H:%M') if group.createdAt else 'Не задано'
                ws['A7'].border = thin_border
                ws['B7'].border = thin_border
                ws['A8'] = "Дата обновления:"
                ws['B8'] = group.updatedAt.strftime('%d.%m.%Y %H:%M') if group.updatedAt else 'Не задано'
                ws['A8'].border = thin_border
                ws['B8'].border = thin_border
                ws['A9'] = "Количество участников:"
                ws['B9'] = len(group.members)
                ws['A9'].border = thin_border
                ws['B9'].border = thin_border
                ws['A10'] = "Количество курсов:"
                ws['B10'] = len(group.courses)
                ws['A10'].border = thin_border
                ws['B10'].border = thin_border

                # Участники
                ws['A12'] = "Участники группы"
                ws['A12'].font = Font(bold=True)
                ws['A12'].border = thin_border
                ws['B12'].border = thin_border
                ws['A13'] = "Имя"
                ws['B13'] = "Логин"
                ws['A13'].font = Font(bold=True)
                ws['B13'].font = Font(bold=True)
                ws['A13'].border = thin_border
                ws['B13'].border = thin_border

                # Заполняем данные участников
                for i, member in enumerate(group.members, start=14):
                    user = member.user
                    ws[f'A{i}'] = self.localize(user.firstName)
                    ws[f'B{i}'] = self.localize(user.username)
                    ws[f'A{i}'].border = thin_border
                    ws[f'B{i}'].border = thin_border

                # Курсы
                start_row = len(group.members) + 16  # Пропускаем строку после участников
                ws[f'A{start_row}'] = "Курсы группы"
                ws[f'A{start_row}'].font = Font(bold=True)
                ws[f'A{start_row}'].border = thin_border
                ws[f'B{start_row}'].border = thin_border
                ws[f'A{start_row+1}'] = "Название курса"
                ws[f'A{start_row+1}'].font = Font(bold=True)
                ws[f'A{start_row+1}'].border = thin_border
                ws[f'B{start_row+1}'].border = thin_border

                # Заполняем данные курсов
                for i, course in enumerate(group.courses, start=start_row+2):
                    title = course['title'] if isinstance(course, dict) and 'title' in course else str(course)
                    ws[f'A{i}'] = self.localize(title)
                    ws[f'A{i}'].border = thin_border
                    ws[f'B{i}'].border = thin_border

                # Выравнивание
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                # Автоматическая настройка ширины столбцов
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        try:
                            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                if column_letter is None:
                                    column_letter = cell.column_letter
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                        except:
                            pass
                    if column_letter:
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column_letter].width = adjusted_width

                # Объединяем ячейки для заголовков разделов
                ws.merge_cells('A1:B1')  # Заголовок
                ws.merge_cells('A3:B3')  # Основная информация
                ws.merge_cells('A12:B12')  # Участники группы
                ws.merge_cells(f'A{start_row}:B{start_row}')  # Курсы группы

                # Сохраняем файл
                file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Excel Files (*.xlsx)")
                if file_path:
                    wb.save(file_path)
                    QMessageBox.information(self, "Успех", "Статистика успешно экспортирована в Excel!")

            except Exception as e:
                raise

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при экспорте в Excel:\n{str(e)}")

    def load_data(self, group_id: int):
        self.group_id = group_id  # Сохраняем ID группы для экспорта
        try:
            group = self.gushub_api.get_group(group_id)
            self.title.setText(f"<h2>Статистика группы: {self.localize(group.name)}</h2>")

            # Основная информация
            info_text = f"""
            <table style='margin-top:0;'>
                <tr><td><b>Название:</b></td><td>{self.localize(group.name)}</td></tr>
                <tr><td><b>Описание:</b></td><td>{self.localize(group.description)}</td></tr>
                <tr><td><b>Код приглашения:</b></td><td>{self.localize(group.inviteCode)}</td></tr>
                <tr><td><b>Дата создания:</b></td><td>{group.createdAt.strftime('%d.%m.%Y %H:%M') if group.createdAt else 'Не задано'}</td></tr>
                <tr><td><b>Дата обновления:</b></td><td>{group.updatedAt.strftime('%d.%m.%Y %H:%M') if group.updatedAt else 'Не задано'}</td></tr>
                <tr><td><b>Количество участников:</b></td><td>{len(group.members)}</td></tr>
                <tr><td><b>Количество курсов:</b></td><td>{len(group.courses)}</td></tr>
            </table>
            """
            self.info_label.setText(info_text)

            # Участники (таблица)
            self._member_ids = []
            if group.members:
                self.members_table.setRowCount(len(group.members))
                for i, m in enumerate(group.members):
                    user = m.user
                    full_name = f"{self.localize(user.firstName)}".strip()
                    self.members_table.setItem(i, 0, QTableWidgetItem(full_name))
                    self.members_table.setItem(i, 1, QTableWidgetItem(self.localize(user.username)))
                    self._member_ids.append(user.id)
                self.members_table.setEnabled(True)
                self.members_table.setStyleSheet("")
            else:
                self.members_table.setRowCount(0)
                self.members_table.setEnabled(False)
                self.members_table.setStyleSheet("QTableWidget { color: #888; font-style: italic; }")
                # Показываем надпись по центру таблицы
                self.members_table.setRowCount(1)
                self.members_table.setSpan(0, 0, 1, 2)
                item = QTableWidgetItem("Нет участников")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                self.members_table.setItem(0, 0, item)
                self._member_ids = []

            # Курсы
            if group.courses:
                courses_text = "<table style='margin-top:0;'>"
                courses_text += "<tr><th>Название курса</th></tr>"
                for c in group.courses:
                    title = c['title'] if isinstance(c, dict) and 'title' in c else str(c)
                    courses_text += f"<tr><td>{self.localize(title)}</td></tr>"
                courses_text += "</table>"
                self.courses_label.setText(courses_text)
                self.courses_label.setAlignment(Qt.AlignmentFlag.AlignTop)
            else:
                self.courses_label.setText("Нет курсов")
                self.courses_label.setAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)

        except Exception as e:
            self.info_label.setText("Ошибка при загрузке данных")
            self.members_table.setRowCount(0)
            self.courses_label.setText("")

    def on_member_selected(self, row: int, column: int):
        # Если нет участников, не реагируем
        if not self.members_table.isEnabled():
            return
        try:
            user_id_item = self.members_table.item(row, 1)
            # Получаем ID студента из исходных данных (лучше хранить список id)
            # Для этого сохраним список id в self._member_ids при загрузке
            student_id = self._member_ids[row]
            self.student_selected.emit(student_id)
        except Exception as e:
            pass

class SelectionWidget(QWidget):
    """Виджет для выбора раздела аналитики"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Создаем горизонтальный layout для кнопок
        buttons_layout = QHBoxLayout()
        buttons_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Добавляем кнопки выбора раздела
        students_button = QPushButton("Список студентов")
        groups_button = QPushButton("Список групп")
        students_button.setFixedWidth(250)
        groups_button.setFixedWidth(250)
        
        # Добавляем кнопки в горизонтальный layout
        buttons_layout.addWidget(students_button)
        buttons_layout.addWidget(groups_button)
        
        # Добавляем горизонтальный layout в основной layout
        layout.addLayout(buttons_layout)
        
        # Подключаем сигналы
        self.students_clicked = students_button.clicked
        self.groups_clicked = groups_button.clicked

class AnalyticsPage(QWidget):
    show_courses_page = pyqtSignal()  # Сигнал для возврата к курсам
    
    def __init__(self):
        super().__init__()
        self.setup_ui()
    
    def setup_ui(self):
        # Создаем основной layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(30)
        
        # Добавляем заголовок
        title_label = QLabel("<h1>Аналитика</h1>")
        title_label.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        layout.addWidget(title_label)
        
        # Создаем основной контент
        content_frame = QFrame()
        content_frame.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Raised)
        content_layout = QVBoxLayout(content_frame)
        content_layout.setContentsMargins(20, 20, 20, 20)
        
        # Создаем стек виджетов для переключения между разделами
        self.stacked_widget = QStackedWidget()
        
        # Создаем виджеты для списков и статистики
        self.selection = SelectionWidget()
        self.students_list = StudentsListWidget()
        self.groups_list = GroupsListWidget()
        self.student_stats = StudentStatsWidget()
        self.group_stats = GroupStatsWidget()
        
        # Подключаем сигналы
        self.selection.students_clicked.connect(self.show_students_list)
        self.selection.groups_clicked.connect(self.show_groups_list)
        self.students_list.student_selected.connect(self.show_student_stats)
        self.groups_list.group_selected.connect(self.show_group_stats)
        self.students_list.back_clicked.connect(self.show_selection)
        self.groups_list.back_clicked.connect(self.show_selection)
        self.student_stats.back_button.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))
        self.group_stats.back_button.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(2))
        self.group_stats.student_selected.connect(self.show_student_stats)
        
        # Добавляем виджеты в стек
        self.stacked_widget.addWidget(self.selection)
        self.stacked_widget.addWidget(self.students_list)
        self.stacked_widget.addWidget(self.groups_list)
        self.stacked_widget.addWidget(self.student_stats)
        self.stacked_widget.addWidget(self.group_stats)
        
        # Добавляем фрейм в основной layout
        layout.addWidget(self.stacked_widget)
        
        # Добавляем кнопку возврата
        return_button = QPushButton("Вернуться к курсам")
        return_button.clicked.connect(self.show_courses_page.emit)
        layout.addWidget(return_button)
    
    def show_selection(self):
        """Показать страницу выбора раздела"""
        self.stacked_widget.setCurrentIndex(0)
    
    def show_students_list(self):
        """Показать список студентов"""
        self.students_list.load_data()
        self.stacked_widget.setCurrentIndex(1)
    
    def show_groups_list(self):
        """Показать список групп"""
        self.groups_list.load_data()
        self.stacked_widget.setCurrentIndex(2)
    
    def show_student_stats(self, student_id: int):
        """Показать статистику студента"""
        self.student_stats.load_data(student_id)
        self.stacked_widget.setCurrentIndex(3)
    
    def show_group_stats(self, group_id: int):
        """Показать статистику группы"""
        self.group_stats.load_data(group_id)
        self.stacked_widget.setCurrentIndex(4)
